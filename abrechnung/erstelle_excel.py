# Erzeugt die Excel-Abrechnung aus berechnung_2026_q2.py
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from berechnung_2026_q2 import rows, gross, rate

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MONTH_NAMES = {"2026-04": "April 2026", "2026-05": "Mai 2026", "2026-06": "Juni 2026"}
MONTHS = ["2026-04", "2026-05", "2026-06"]
OBJECTS = ["Studio", "Apartment"]

wb = Workbook()

BLUE = "1F4E5F"
LIGHT = "EAF1F4"
head_font = Font(bold=True, color="FFFFFF")
head_fill = PatternFill("solid", fgColor=BLUE)
sub_fill = PatternFill("solid", fgColor=LIGHT)
bold = Font(bold=True)
thin = Side(style="thin", color="BBBBBB")
border = Border(bottom=thin)
chf = '#,##0.00'

# ---------- Blatt 1: Übersicht ----------
ws = wb.active
ws.title = "Übersicht"
ws["A1"] = "Provisionsabrechnung Ferienwohnungs-Management"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Kunde: KaffeeKlatsch Wallis AG — B&B Brig by KaffeeKlatsch (Studio & Apartment)"
ws["A3"] = "Periode: April – Juni 2026 | Portale: Booking.com & Airbnb | Währung: CHF"

headers = ["Monat", "Objekt", "Buchungen", "Umsatz brutto", "Portal-Kommission",
           "Reinigungspauschalen", "Provisionsbasis", "Satz", "Provision"]
r0 = 5
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=r0, column=c, value=h)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = r0 + 1
tot = dict(n=0, g=0.0, k=0.0, p=0.0, b=0.0, prov=0.0)
for m in MONTHS:
    for obj in OBJECTS:
        rs = [x for x in rows if x["monat"] == m and x["objekt"] == obj]
        if not rs:
            continue
        n = len(rs)
        g = round(sum(x["brutto"] for x in rs), 2)
        k = round(sum(x["kommission"] for x in rs), 2)
        p = round(sum(x["pauschale"] for x in rs), 2)
        b = round(sum(x["basis"] for x in rs), 2)
        prov = round(sum(x["provision"] for x in rs), 2)
        vals = [MONTH_NAMES[m], obj, n, g, k, p, b, rate[(m, obj)], prov]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c in (4, 5, 6, 7, 9):
                cell.number_format = chf
            if c == 8:
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal="center")
        tot["n"] += n; tot["g"] += g; tot["k"] += k; tot["p"] += p; tot["b"] += b; tot["prov"] += prov
        r += 1
    # Monats-Zwischensumme
    rs_m = [x for x in rows if x["monat"] == m]
    ws.cell(row=r, column=1, value=f"Total {MONTH_NAMES[m]}").font = bold
    ws.cell(row=r, column=3, value=len(rs_m)).font = bold
    for c, key in ((4, "brutto"), (5, "kommission"), (6, "pauschale"), (7, "basis"), (9, "provision")):
        cell = ws.cell(row=r, column=c, value=round(sum(x[key] for x in rs_m), 2))
        cell.font = bold
        cell.number_format = chf
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = sub_fill
    r += 1

r += 1
ws.cell(row=r, column=1, value="TOTAL Provision April – Juni 2026").font = Font(bold=True, size=12)
cell = ws.cell(row=r, column=9, value=round(tot["prov"], 2))
cell.font = Font(bold=True, size=12)
cell.number_format = chf

r += 2
notes = [
    "Berechnungsregeln:",
    "• Provisionssatz pro Objekt und Monat: Brutto-Monatsumsatz (vor Portal-Kommission) bis CHF 1'600 → 5 %, über CHF 1'600 → 10 %.",
    "• Provisionsbasis pro Buchung: Buchungswert ./. Portal-Kommission ./. Reinigungspauschale (Studio CHF 100 / Apartment CHF 130).",
    "• Airbnb zahlt netto aus (Servicegebühr bereits abgezogen); für den Schwellenwert zählen die Bruttoeinkünfte.",
    "• Monatszuordnung nach Abreisedatum (entspricht der Rechnungsperiode von Booking.com).",
    "• Stornierte Buchungen (4× Booking.com, Betrag CHF 0) sind nicht berücksichtigt.",
    "• Buchung Kolb (Abreise 01.04.) bereits mit März abgerechnet; Buchung Passant (Abreise 03.07.) folgt mit Juli.",
    "• Beträge exkl. allfälliger MwSt.",
]
for note in notes:
    ws.cell(row=r, column=1, value=note)
    r += 1

widths = [14, 12, 11, 14, 17, 20, 15, 8, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Blatt 2: Details ----------
ws2 = wb.create_sheet("Details pro Buchung")
headers2 = ["Monat", "Objekt", "Portal", "Gast", "Aufenthalt", "Buchungswert brutto",
            "Portal-Kommission", "Netto nach Kommission", "Reinigungspauschale",
            "Provisionsbasis", "Satz", "Provision"]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = 2
for m in MONTHS:
    for obj in OBJECTS:
        rs = [x for x in rows if x["monat"] == m and x["objekt"] == obj]
        for x in rs:
            vals = [MONTH_NAMES[m], obj, x["portal"], x["gast"], x["zeitraum"], x["brutto"],
                    x["kommission"], x["netto"], x["pauschale"], x["basis"], x["satz"], x["provision"]]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = border
                if c in (6, 7, 8, 9, 10, 12):
                    cell.number_format = chf
                if c == 11:
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal="center")
            r += 1
        if rs:
            ws2.cell(row=r, column=4, value=f"Zwischensumme {MONTH_NAMES[m]} – {obj} "
                     f"(Umsatz {gross[(m,obj)]:,.2f} → {int(rate[(m,obj)]*100)} %)").font = bold
            for c, key in ((6, "brutto"), (7, "kommission"), (9, "pauschale"), (10, "basis"), (12, "provision")):
                cell = ws2.cell(row=r, column=c, value=round(sum(x[key] for x in rs), 2))
                cell.font = bold
                cell.number_format = chf
            for c in range(1, 13):
                ws2.cell(row=r, column=c).fill = sub_fill
            r += 1

r += 1
ws2.cell(row=r, column=4, value="TOTAL Provision April – Juni 2026").font = Font(bold=True, size=12)
cell = ws2.cell(row=r, column=12, value=round(sum(x["provision"] for x in rows), 2))
cell.font = Font(bold=True, size=12)
cell.number_format = chf

widths2 = [12, 11, 12, 28, 15, 13, 13, 13, 13, 13, 7, 11]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Provisionsabrechnung_KaffeeKlatsch_2026-04_bis_2026-06.xlsx")
wb.save(out)
print("geschrieben:", out)
